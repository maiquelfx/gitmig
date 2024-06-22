!pip install GitPython
!apt-get install graphviz
!pip install git-graph
import shutil
import os
import regex as re
import pandas as pd
from datetime import datetime
from git import Repo, GitCommandError
import numpy as np
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl
from copy import copy
import zipfile
from google.colab import files
from google.colab import files as colab_files
import subprocess
import requests


groups = {
        'MANAGEMENT': [ #gerenciamento / documentação
            "explain", "readme", "section", "directory", "tidy", "docs", "page", "adopt", "bump", "changelog", "clean", "code review", "comment",
            "copyright", "documentation", "document", "format", "formatting", "dependencies","resume",
            "integrat", "javadoc", "license", "maintenance", "manual", "merge",
            "organiz", "polish", "readme", "repository", "docker",
            "structure", "style", "TODO", "upgrade", "install", "pystan", "describe", "require", "note", "typo", "spelling", "pep8", "installation",
            "recommend", "set", "requirements", "gitignore", "git","instructions", "translate", "guide", "updating", "auto", "specify", "annotated",
            "MD", ".md", "ignoring", "ignore", "numbering", "merging", "tweak", "mark dev", "mark version"
        ],
        'CORRECTIVE_ENGINEERING': [ #correção,
            "broke", "bug", "appease", "send", "correct", "deprecat", "error", "fixed", "fix", "handle",
            "harmonize", "issue", "kill", "penalty", "proper", "reduce", "repair", "revise",
            "avoid", "convert", "clear","custom", "need", "compatible", "conflicts", "reversed", "wrong", "exception", "go back", "workaround", #gambiarra
            "resolve", "put"

        ],
        'FORWARD_ENGINEERING': [ #novas funcionalidades ou melhorias #engenharia incremental
            "add", "added", "wip", "convert", "safely", "display", "translation", "api", "solution",
            "hotkeys", "Buffering", "separate", "adding", "allow", "anchor", "build", "button", "cache",
            "catch", "check", "command", "compiler", "completed", "configurable",
            "create", "created", "creat", "development", "enablement", "filter", "finish", "first",
            "functionality", "gateway", "implemented", "implement", "increas",
            "inference", "initial", "introduc", "log", "logging", "new", "prevent",
            "progress", "provide", "refresh", "register", "rename", "render", "request",
            "return", "save", "schedule", "script", "server", "show", "split",
            "start", "statistics", "store", "support", "swapped", "template",
            "tracking", "train", "use", "worker", "switch", "arguments", "feat", "instruct",
            "seasonality", "stan model", "trends", "wheels", "expose", "raise", "regressors", "transition", "scrape",
            "prototype", "path", "alternative approach"
        ],
        'REENGINEERING': [ #refatoração #otimização
            "adjust", "migrat", "switch", "better", "best", "blocked", "chang", "combine", "consolidate", "decreased", "speed"
            "delete", "del", "deleting", "design", "change", "duplicat", "eliminat", "enhance", "enhanced",
            "extend", "import", "improv", "made", "make", "miss", "modif", "move", "force",
            "moved", "obsolete", "optimiz", "order", "override", "refactor", "wrapped", "less", "wire",
            "rearrang", "replac", "reorder", "reorganiz", "replac", "restructure", "prefer", "removing", "don't repeat",
            "revert", "rewrote", "simplif", "sort", "tinkering", "update", "every", "dry", "refin", # DRY (Don't Repeat Yourself)
            "regenerate", "styling", "speed", "optimize", "encode", "enabl", "disabl", "compatibility", "removal", "purification", "->", "overwrite", "Strip",
            "rised", "adapted", "instead", "rewritten", "standart", "standardization", "yamlized", "reorganising", "re-", "re-*", "redefine","edit",
            "backtrack"
        ],
        'TESTS': [
            "assert", "eval", "example", "examples", "test", "tests", "verify",
            "unittest", "pytest", "check", "validate", "validation", "travis", "run"     #TRAVIS ferramenta de IC/CD
        ],
        'RELEASES': [ "release",
        r"(?i)release\s*[\d.]+",  # Captura "release x.x.x"
        r"(?i)version\s*[\d.]+",  # Captura "version x.x.x"
        r"(?i)v?\d+\.\d+\.\d+(?:[-\w\d]*\.\d+)?",  # Captura "v1.2.3", "1.2.3", "1.2.3-alpha.4"
    ]
    }

extension_to_language = {
      'py': 'Python',
      'cpp': 'C++',
      'h': 'C++',
      'hpp': 'C++',
      'java': 'Java',
      'js': 'JavaScript',
      'html': 'HTML',
      'htm': 'HTML',
      'css': 'CSS',
      'php': 'PHP',
      'rb': 'Ruby',
      'cs': 'C#',
      'swift': 'Swift',
      'go': 'Go',
      'yml': 'YAML',
      'ts': 'TypeScript',
      'sh': 'Shell Script',
      'bash': 'Shell Script',
      'pl': 'Perl',
      'kt': 'Kotlin',
      'rs': 'Rust',
      'm': 'Objective-C',
      'scala': 'Scala',
      'lua': 'Lua',
      'r': 'R',
      'm': 'MATLAB',
      'xml': 'XML'
  }

def prospect(): #apenas para prospecção git
  token = 'ghp_token'
  headers = {'Authorization': f'token {token}'} if token else {}

  def search_repos(query, max_repos=1000):
      url = f'https://api.github.com/search/repositories?q={query}&per_page=100'
      repos = []
      while url and len(repos) < max_repos:
          response = requests.get(url, headers=headers)
          if response.status_code != 200:
              raise Exception(f'Error searching repos: {response.status_code}')
          repos.extend(response.json().get('items', []))
          if 'next' in response.links:
              url = response.links['next']['url']
          else:
              url = None
      return repos[:max_repos]

  def get_commit_count(repo):
      url = f'https://api.github.com/repos/{repo}/commits?per_page=1'
      response = requests.get(url, headers=headers)
      if response.status_code != 200:
          raise Exception(f'Error fetching commit count for {repo}: {response.status_code}')
      if 'Link' in response.headers:
          link_header = response.headers['Link']
          commit_count = int(link_header.split(',')[1].split('&page=')[1].split('>')[0])
      else:
          commit_count = 1  # obs: se houver apenas um commit, o cabeçalho de link não estará presente
      return commit_count

  def list_repos_with_commits_in_range(query, min_commits=500, max_commits=1000, max_repos=1000):
      repos = search_repos(query, max_repos)
      eligible_repos = []
      for repo in repos:
          repo_name = repo['full_name']
          try:
              commit_count = get_commit_count(repo_name)
              if min_commits <= commit_count <= max_commits:
                  eligible_repos.append((repo_name, commit_count))
                  print(f"{repo_name}: {commit_count} commits")
          except Exception as e:
              print(f"Error processing {repo_name}: {e}")
      return eligible_repos

  # searching
  query = f'stars:>=500 language:Python'  # repositórios de interesse
  min_commits = 500
  max_commits = 1000
  repos_in_range = list_repos_with_commits_in_range(query, min_commits, max_commits)

  print(f"Total repositories with between {min_commits} and {max_commits} commits: {len(repos_in_range)}")




def repo():
    global link, nome, user #serão usados em outros pontos 
    input_str = input("Digite o caminho: ")
    parts = input_str.split("/")
    if len(parts) == 2 and parts[0] and parts[1]:
        link = input_str
        nome = parts[1]  # Armazenar o nome do repositório
        user = parts[0] # Armazenar o nome do proprietário
        print("Projeto Atual:", link)
    else:
        print("Entrada inválida. Certifique-se de que o formato seja 'user/repo'.")

def clone():
    global link
    if 'link' not in globals():
        print("Nenhum link foi definido. Use a função 'repo()' para definir um link primeiro.")
        return
    os.system(f"git clone https://github.com/{link}")

def mining():
    global nome
    global arquivo
    if 'nome' not in globals():
        print("O nome do repositório não está definido. Por favor, defina-o primeiro usando a função 'repo()'.")
        return

    # Nome do arquivo de saída
    arquivo = f"{nome}.xlsx"
    repo_path = f"{nome}"
    repo = Repo(repo_path)
    # Obter todos os commits do repositório
    commits = list(repo.iter_commits())
    commit_data = []
    for commit in commits:
        author_name = commit.author.name
        author_email = commit.author.email
        commit_date = commit.committed_datetime
        commit_message = commit.message
        files_changed = len(commit.stats.files)
        branch_name = None
        for branch in repo.branches:
            if commit in branch.commit.iter_parents():
                branch_name = branch.name
                break
        commit_data.append({
            'SHA': commit.hexsha,
            'author': author_name,
            'Email': author_email,
            'Data_Hora': commit_date,
            'message': commit_message,
            'Amt_files': files_changed,
            'branch_name': branch_name
        })

    df = pd.DataFrame(commit_data)
    df.columns = ['SHA', 'author', 'Email', 'Data_Hora', 'message', 'Amt_files', 'branch_name']
    df['Date'] = df['Data_Hora'].apply(lambda x: x.strftime("%Y-%m-%d"))
    df['Hour'] = df['Data_Hora'].apply(lambda x: x.strftime("%H:%M:%S"))
    df.drop(columns=['Data_Hora'], inplace=True)
    df['Day'] = df['Date'].apply(lambda x: datetime.strptime(x, "%Y-%m-%d").strftime("%A"))
    df['Short_Message'] = df['message'].apply(lambda x: x.split('\n')[0])
    df['Full_Message'] = df['message'].apply(lambda x: '\n'.join(x.split('\n')[1:]))

    # Reorganizar 
    df = df[['Amt_files', 'SHA', 'author', 'Email', 'Date', 'Day', 'Hour', 'Short_Message', 'Full_Message', 'branch_name']]
    df.to_excel(arquivo, index=False)
    print(f"Os dados do repositório '{nome}' foram minerados e salvos em '{arquivo}'.")

def cat(entrada, saida): #disc.: only base
    def classify_commit(short_message, full_message):
        for category, keywords in groups.items():
            for word in keywords:
                if word in short_message.lower() or word in full_message.lower():
                    return category
        return 'UNCLASSIFIED'

    df = pd.read_excel(entrada)
    short_message_col = [col for col in df.columns if 'Short' in col][0]
    full_message_col = [col for col in df.columns if 'Full' in col][0]
    #add class
    df['Classification'] = df.apply(lambda row: classify_commit(str(row[short_message_col]), str(row[full_message_col])), axis=1)
    df.to_excel(saida, index=False)


#################

def cat_color(entrada, saida):
    colors = {
        'MANAGEMENT': 'FFFF99',  # Light Yellow
        'CORRECTIVE_ENGINEERING': 'FF9999',  # Light Red
        'FORWARD_ENGINEERING': '99FF99',  # Light Green
        'REENGINEERING': '99CCFF',  # Light Blue
        'TESTS': 'FFCC99',  # Light Orange
        'UNCLASSIFIED': 'FFFFFF',  # White
        'RELEASES': '9999FF'  # Light Purple
    }

    # Classificação
    def classify_commit(short_message, full_message):
        for category, keywords in groups.items():
            for pattern in keywords:
                if re.search(pattern, short_message.lower()) or re.search(pattern, full_message.lower()):
                    return category
        return 'UNCLASSIFIED'
    df = pd.read_excel(entrada)
    short_message_col = [col for col in df.columns if 'Short' in col][0]
    full_message_col = [col for col in df.columns if 'Full' in col][0]
    df['Classification'] = df.apply(lambda row: classify_commit(str(row[short_message_col]), str(row[full_message_col])), axis=1)
    df.to_excel(saida, index=False)
    wb = load_workbook(saida)
    ws = wb.active

    classification_col = None
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == 'Classification':
            classification_col = col[0].column_letter
            break

    if classification_col is None:
        raise ValueError("A coluna 'Classification' não foi encontrada no arquivo Excel.")

    for row in range(2, ws.max_row + 1):  # Começa em 2 para ignorar o cabeçalho
        classification = ws[f'{classification_col}{row}'].value
        fill_color = colors.get(classification, 'FFFFFF')
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for cell in ws[row]:
            cell.fill = fill
    wb.save(saida)



################

def cat_class():
  saida = nome + "_class_.xlsx"
  entrada = f'{nome}_class.xlsx'
  df = pd.read_excel(f'{nome}_class.xlsx')

  # Ordenar Z a A
  df_sorted = df.sort_values(by='Classification', ascending=False)
  df_sorted = pd.concat([df_sorted[df_sorted['Classification'] == 'UNCLASSIFIED'],
                        df_sorted[df_sorted['Classification'] != 'UNCLASSIFIED']])

  df_sorted.to_excel(saida, index=False)

  print(df_sorted)

''' disc.: att
def cat_class_color():
  saida = nome + "_class_color_.xlsx"
  entrada = f'{nome}_class_color.xlsx'
  df = pd.read_excel(f'{nome}_class_color.xlsx')

  df_sorted = df.sort_values(by='Classification', ascending=False)

  df_sorted = pd.concat([df_sorted[df_sorted['Classification'] == 'UNCLASSIFIED'],
                        df_sorted[df_sorted['Classification'] != 'UNCLASSIFIED']])

  df_sorted.to_excel(saida, index=False)

  print(df_sorted)
'''

def cat_class_color():
    saida = nome + "_class_color_sorted.xlsx"
    entrada = f'{nome}_class_color.xlsx'

    workbook = openpyxl.load_workbook(filename=entrada)
    worksheet = workbook.active

    data = [(copy(worksheet.cell(row=row, column=worksheet.max_column).value),
             [copy(cell.value) for cell in worksheet[row]],
             copy(worksheet.cell(row=row, column=1).fill.start_color.rgb)) for row in range(2, worksheet.max_row + 1)]

    data.sort(key=lambda x: x[0] if x[0] != 'UNCLASSIFIED' else '', reverse=False) # true // false para reverter

    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active

    for col in range(1, worksheet.max_column + 1):
        new_worksheet.cell(row=1, column=col, value=worksheet.cell(row=1, column=col).value)

    for row, (classification, values, fill_color) in enumerate(data, start=2):
        for col, value in enumerate(values, start=1):
            new_worksheet.cell(row=row, column=col, value=value)
            new_worksheet.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    new_workbook.save(saida)

######################

def cat_shuffle():
    entrada = f'{nome}_class.xlsx'
    saida = f'shuffled_{nome}_class.xlsx'
    df = pd.read_excel(entrada)
    def intercalate_categories(df):
        # Obter as categorias únicas
        unique_categories = df['Classification'].unique()
        np.random.shuffle(unique_categories)
        category_dfs = [df[df['Classification'] == category].reset_index(drop=True) for category in unique_categories]
        max_length = max(len(category_df) for category_df in category_dfs)
        intercalated_rows = []

        for i in range(max_length):
            for category_df in category_dfs:
                if i < len(category_df):
                    intercalated_rows.append(category_df.iloc[i])

        intercalated_df = pd.DataFrame(intercalated_rows).reset_index(drop=True)
        return intercalated_df

    shuffled_df = intercalate_categories(df)
    shuffled_df.to_excel(saida, index=False)
    print("Arquivo salvo como 'shuffled.xlsx'")


def cat_shuffle_color():
    entrada = f'{nome}_class_color.xlsx'
    saida = f'shuffled_{nome}_class_color.xlsx'
    df = pd.read_excel(entrada)
    def intercalate_categories(df):
        unique_categories = df['Classification'].unique()
        np.random.shuffle(unique_categories)
        category_dfs = [df[df['Classification'] == category].reset_index(drop=True) for category in unique_categories]
        max_length = max(len(category_df) for category_df in category_dfs)
        intercalated_rows = []

        for i in range(max_length):
            for category_df in category_dfs:
                if i < len(category_df):
                    intercalated_rows.append(category_df.iloc[i])
        intercalated_df = pd.DataFrame(intercalated_rows).reset_index(drop=True)
        return intercalated_df
    shuffled_df = intercalate_categories(df)
    shuffled_df.to_excel(saida, index=False)
    print("Arquivo salvo 'file-shuffled.xlsx'")


def data():
  print(f"Repositório: {nome}")
  print()

  def get_directory_size(directory):
      total_size = 0
      for dirpath, dirnames, filenames in os.walk(directory):
          for f in filenames:
              fp = os.path.join(dirpath, f)
              if os.path.isfile(fp):
                  total_size += os.path.getsize(fp)
      return total_size
  current_path = os.getcwd()
  repo_path = os.path.join(current_path, f'{nome}')
  size = get_directory_size(repo_path)
  print(f"Tamanho do repositório: {size / 1024 / 1024:.2f} MB")
  print(f'{user}/{nome}')
  df = pd.read_excel(f'{nome}_class.xlsx')
  commit_counts_per_group = df['Classification'].value_counts()
  print()
  print("Quantidade de commits por grupo:")
  print(commit_counts_per_group)
  total_commits = len(df)
  print("\nQuantidade total de commits:", total_commits)

  # Qual usuário que mais fez commits?
  most_active_user = df['author'].value_counts().idxmax()
  num_commits_most_active_user = df['author'].value_counts().max()

  print(f"\nUsuário que mais fez commits: {most_active_user}")
  print(f"Quantidade de commits realizados pelo usuário: {num_commits_most_active_user}")

  # Quantidade de usuários
  num_users = df['author'].nunique()
  print(f"\nQuantidade de usuários: {num_users}")

  # Quais os três dias da semana com mais commits?
  df['Date'] = pd.to_datetime(df['Date'])
  df_weekly_commits = df.groupby(df['Date'].dt.strftime('%A')).size().nlargest(3)
  print("\nTrês dias da semana com mais commits:")
  print(df_weekly_commits)

  # Qual dia do projeto teve menos commits?
  df['Dia_Semana'] = df['Date'].dt.strftime('%A')
  commits_por_dia = df['Dia_Semana'].value_counts()
  dia_menos_commits = commits_por_dia.idxmin()
  num_commits_menos = commits_por_dia.min()

  print(f"\nDia do projeto com menos commits: {dia_menos_commits}")
  print(f"Quantidade de commits realizados nesse dia: {num_commits_menos}")

  # Métricas por grupo
  print('\n###################')

  # Probabilidade de ocorrência
  total_commits = commit_counts_per_group.sum()
  prob_per_group = commit_counts_per_group / total_commits

  # Mostrar a probabilidade de ocorrência
  print("Probabilidade de ocorrência:")
  print(prob_per_group)

  media = prob_per_group.mean().round(20)
  variancia = prob_per_group.var()
  desvio_padrao = prob_per_group.std()
  print("\nMétricas por grupo:")
  print("Média:", media.round(20))
  print("Variância:", variancia)

  # >> desvio padrão para cada subgrupo
  media = prob_per_group.mean()
  desvios_padrao_por_grupo = []

  for group, probabilidade in prob_per_group.items():
      variancia = (probabilidade - media) ** 2
      desvio_padrao_grupo = np.sqrt(variancia)
      desvios_padrao_por_grupo.append((group, desvio_padrao_grupo))
  print("\nDesvio Padrão por grupo:")
  for grupo, desvio_padrao in desvios_padrao_por_grupo:
      print(f"{grupo}: {desvio_padrao}")


  print()
  print()
  print('################====UNCLASSIFIED=====####################')

  print()
  print()

  unclassified_commits = df[df['Classification'] == 'UNCLASSIFIED'].copy()
  num_unclassified_commits = len(unclassified_commits)
  print(f"\nNúmero de commits 'UNCLASSIFIED': {num_unclassified_commits}")

  if num_unclassified_commits > 0:
      most_active_unclassified_user = unclassified_commits['author'].value_counts().idxmax()
      num_commits_most_active_unclassified_user = unclassified_commits['author'].value_counts().max()

      print(f"\nUsuário que mais fez commits 'UNCLASSIFIED': {most_active_unclassified_user}")
      print(f"Quantidade de commits 'UNCLASSIFIED' realizados pelo usuário: {num_commits_most_active_unclassified_user}")

      unclassified_commits['Date'] = pd.to_datetime(unclassified_commits['Date'])
      unclassified_weekly_commits = unclassified_commits.groupby(unclassified_commits['Date'].dt.day_name()).size().nlargest(3)
      print("\nTrês dias da semana com mais commits 'UNCLASSIFIED':")
      print(unclassified_weekly_commits)

      unclassified_commits['Dia_Semana'] = unclassified_commits['Date'].dt.strftime('%A')
      commits_unclassified_por_dia = unclassified_commits['Dia_Semana'].value_counts()
      dia_menos_unclassified_commits = commits_unclassified_por_dia.idxmin()
      num_commits_menos_unclassified = commits_unclassified_por_dia.min()

      print(f"\nDia com menos commits 'UNCLASSIFIED': {dia_menos_unclassified_commits}")
      print(f"Quantidade de commits 'UNCLASSIFIED' realizados nesse dia: {num_commits_menos_unclassified}")

      max_files_changed_commit = unclassified_commits.loc[unclassified_commits['Amt_files'].idxmax()]
      user_max_files_changed = max_files_changed_commit['author']
      max_files_changed = max_files_changed_commit['Amt_files']

      print(f"\nUsuário que fez o commit com maior número de alterações em arquivos: {user_max_files_changed}")
      print(f"Quantidade de arquivos alterados: {max_files_changed}")

      print("======================================================================")
      max_files_changed_commit = unclassified_commits.loc[unclassified_commits['Amt_files'].idxmax()]
      user_max_files_changed = max_files_changed_commit['author']
      max_files_changed = max_files_changed_commit['Amt_files']
      commit_message_short = max_files_changed_commit['Short_Message']
      commit_message_full = max_files_changed_commit['Full_Message']
      print(f"Mensagem do commit (Short): {commit_message_short}")
      print(f"Mensagem do commit (Full): {commit_message_full}")
  else:
      print("\nNão há commits 'UNCLASSIFIED' para analisar.")



def data_color():
  print(f"Repositório: {nome}")
  print()

  def get_directory_size(directory):
      total_size = 0
      for dirpath, dirnames, filenames in os.walk(directory):
          for f in filenames:
              fp = os.path.join(dirpath, f)
              if os.path.isfile(fp):
                  total_size += os.path.getsize(fp)
      return total_size

  current_path = os.getcwd()
  repo_path = os.path.join(current_path, f'{nome}')
  size = get_directory_size(repo_path)
  print(f'{user}/{nome}')
  print(f"Tamanho do repositório: {size / 1024 / 1024:.2f} MB")
  df = pd.read_excel(f'{nome}_class_color.xlsx')
  commit_counts_per_group = df['Classification'].value_counts()
  print()
  print("Quantidade de commits por grupo:")
  print(commit_counts_per_group)
  total_commits = len(df)
  print("\nQuantidade total de commits:", total_commits)

  # Qual usuário que mais fez commits?
  most_active_user = df['author'].value_counts().idxmax()
  num_commits_most_active_user = df['author'].value_counts().max()

  print(f"\nUsuário que mais fez commits: {most_active_user}")
  print(f"Quantidade de commits realizados pelo usuário: {num_commits_most_active_user}")

  num_users = df['author'].nunique()
  print(f"\nQuantidade de usuários: {num_users}")

  df['Date'] = pd.to_datetime(df['Date'])
  df_weekly_commits = df.groupby(df['Date'].dt.strftime('%A')).size().nlargest(3)
  print("\nTrês dias da semana com mais commits:")
  print(df_weekly_commits)
  df['Dia_Semana'] = df['Date'].dt.strftime('%A')
  commits_por_dia = df['Dia_Semana'].value_counts()
  dia_menos_commits = commits_por_dia.idxmin()
  num_commits_menos = commits_por_dia.min()

  print(f"\nDia do projeto com menos commits: {dia_menos_commits}")
  print(f"Quantidade de commits realizados nesse dia: {num_commits_menos}")

  # Métricas por grupo
  print('\n###################')

  total_commits = commit_counts_per_group.sum()
  prob_per_group = commit_counts_per_group / total_commits

  print("Probabilidade de ocorrência:")
  print(prob_per_group)

  media = prob_per_group.mean().round(20)
  variancia = prob_per_group.var()
  desvio_padrao = prob_per_group.std()

  # Mostrar as métricas
  print("\nMétricas por grupo:")
  print("Média:", media.round(20))
  print("Variância:", variancia)

  media = prob_per_group.mean()
  desvios_padrao_por_grupo = []

  for group, probabilidade in prob_per_group.items():
      variancia = (probabilidade - media) ** 2
      desvio_padrao_grupo = np.sqrt(variancia)
      desvios_padrao_por_grupo.append((group, desvio_padrao_grupo))

  print("\nDesvio Padrão por grupo:")
  for grupo, desvio_padrao in desvios_padrao_por_grupo:
      print(f"{grupo}: {desvio_padrao}")


  print()
  print()
  print('################====UNCLASSIFIED=====####################')

  print()
  print()

  unclassified_commits = df[df['Classification'] == 'UNCLASSIFIED'].copy()

  num_unclassified_commits = len(unclassified_commits)
  print(f"\nNúmero de commits 'UNCLASSIFIED': {num_unclassified_commits}")

  if num_unclassified_commits > 0:
      most_active_unclassified_user = unclassified_commits['author'].value_counts().idxmax()
      num_commits_most_active_unclassified_user = unclassified_commits['author'].value_counts().max()

      print(f"\nUsuário que mais fez commits 'UNCLASSIFIED': {most_active_unclassified_user}")
      print(f"Quantidade de commits 'UNCLASSIFIED' realizados pelo usuário: {num_commits_most_active_unclassified_user}")

      # Datas dos commits "UNCLASSIFIED"
      unclassified_commits['Date'] = pd.to_datetime(unclassified_commits['Date'])
      unclassified_weekly_commits = unclassified_commits.groupby(unclassified_commits['Date'].dt.day_name()).size().nlargest(3)
      print("\nTrês dias da semana com mais commits 'UNCLASSIFIED':")
      print(unclassified_weekly_commits)

      # Dia com menos commits "UNCLASSIFIED"
      unclassified_commits['Dia_Semana'] = unclassified_commits['Date'].dt.strftime('%A')
      commits_unclassified_por_dia = unclassified_commits['Dia_Semana'].value_counts()
      dia_menos_unclassified_commits = commits_unclassified_por_dia.idxmin()
      num_commits_menos_unclassified = commits_unclassified_por_dia.min()

      print(f"\nDia com menos commits 'UNCLASSIFIED': {dia_menos_unclassified_commits}")
      print(f"Quantidade de commits 'UNCLASSIFIED' realizados nesse dia: {num_commits_menos_unclassified}")

      # Usuário que fez o commit com maior número de alterações em arquivos
      max_files_changed_commit = unclassified_commits.loc[unclassified_commits['Amt_files'].idxmax()]
      user_max_files_changed = max_files_changed_commit['author']
      max_files_changed = max_files_changed_commit['Amt_files']

      print(f"\nUsuário que fez o commit com maior número de alterações em arquivos: {user_max_files_changed}")
      print(f"Quantidade de arquivos alterados: {max_files_changed}")

      print("======================================================================")
      # Usuário que fez o commit com maior número de alterações em arquivos
      max_files_changed_commit = unclassified_commits.loc[unclassified_commits['Amt_files'].idxmax()]
      user_max_files_changed = max_files_changed_commit['author']
      max_files_changed = max_files_changed_commit['Amt_files']
      commit_message_short = max_files_changed_commit['Short_Message']
      commit_message_full = max_files_changed_commit['Full_Message']
      print(f"Mensagem do commit (Short): {commit_message_short}")
      print(f"Mensagem do commit (Full): {commit_message_full}")
  else:
      print("\nNão há commits 'UNCLASSIFIED' para analisar.")


def data_color_txt(nome):
  with open(f'{nome}_output.txt', 'w') as f:
        def print_to_file(*args, **kwargs):
            print(*args, **kwargs)
            print(*args, **kwargs, file=f)

        print_to_file(f"Repositório: {nome}")
        print_to_file()

        def get_directory_size(directory):
            total_size = 0
            for dirpath, dirnames, filenames in os.walk(directory):
                for file in filenames:
                    fp = os.path.join(dirpath, file)
                    if os.path.isfile(fp):
                        total_size += os.path.getsize(fp)
            return total_size

        current_path = os.getcwd()
        repo_path = os.path.join(current_path, f'{nome}')
        size = get_directory_size(repo_path)
        print(f'{user}/{nome}')
        print_to_file(f"Tamanho do repositório: {size / 1024 / 1024:.2f} MB")
        df = pd.read_excel(f'{nome}_class_color.xlsx')
        commit_counts_per_group = df['Classification'].value_counts()
        print_to_file()
        print_to_file("Quantidade de commits por grupo:")
        print_to_file(commit_counts_per_group)
        total_commits = len(df)
        print_to_file("\nQuantidade total de commits:", total_commits)
        most_active_user = df['author'].value_counts().idxmax()
        num_commits_most_active_user = df['author'].value_counts().max()

        print_to_file(f"\nUsuário que mais fez commits: {most_active_user}")
        print_to_file(f"Quantidade de commits realizados pelo usuário: {num_commits_most_active_user}")

        # Quantidade de usuários
        num_users = df['author'].nunique()
        print_to_file(f"\nQuantidade de usuários: {num_users}")

        df['Date'] = pd.to_datetime(df['Date'])
        df_weekly_commits = df.groupby(df['Date'].dt.strftime('%A')).size().nlargest(3)
        print_to_file("\nTrês dias da semana com mais commits:")
        print_to_file(df_weekly_commits)

        df['Dia_Semana'] = df['Date'].dt.strftime('%A')
        commits_por_dia = df['Dia_Semana'].value_counts()
        dia_menos_commits = commits_por_dia.idxmin()
        num_commits_menos = commits_por_dia.min()

        print_to_file(f"\nDia do projeto com menos commits: {dia_menos_commits}")
        print_to_file(f"Quantidade de commits realizados nesse dia: {num_commits_menos}")

        # Métricas por grupo
        print_to_file('\n###################')

        total_commits = commit_counts_per_group.sum()
        prob_per_group = commit_counts_per_group / total_commits

        print_to_file("Probabilidade de ocorrência:")
        print_to_file(prob_per_group)      
        classifications = list(prob_per_group.index)
        probabilities = list(prob_per_group.values)

        classifications.insert(0, nome)
        probabilities.insert(0, "")
        metrics_df = pd.DataFrame({
            'Classification': classifications,
            'Probabilidade': probabilities
        })

        metrics_file = f'{nome}_metrica.xlsx'
        metrics_df.to_excel(metrics_file, index=False)

        print_to_file('\nMétricas salvas em:', metrics_file)
        media = prob_per_group.mean().round(20)
        variancia = prob_per_group.var()
        desvio_padrao = prob_per_group.std()
        print_to_file("\nMétricas por grupo:")
        print_to_file("Média:", media.round(20))
        print_to_file("Variância:", variancia)
        media = prob_per_group.mean()
        desvios_padrao_por_grupo = []

        for group, probabilidade in prob_per_group.items():
            variancia = (probabilidade - media) ** 2
            desvio_padrao_grupo = np.sqrt(variancia)
            desvios_padrao_por_grupo.append((group, desvio_padrao_grupo))
        print_to_file("\nDesvio Padrão por grupo:")
        for grupo, desvio_padrao in desvios_padrao_por_grupo:
            print_to_file(f"{grupo}: {desvio_padrao}")

        print_to_file()
        print_to_file()
        print_to_file('################====UNCLASSIFIED=====####################')
        print_to_file()
        print_to_file()
        unclassified_commits = df[df['Classification'] == 'UNCLASSIFIED'].copy()
        num_unclassified_commits = len(unclassified_commits)
        print_to_file(f"\nNúmero de commits 'UNCLASSIFIED': {num_unclassified_commits}")

        if num_unclassified_commits > 0:
            most_active_unclassified_user = unclassified_commits['author'].value_counts().idxmax()
            num_commits_most_active_unclassified_user = unclassified_commits['author'].value_counts().max()

            print_to_file(f"\nUsuário que mais fez commits 'UNCLASSIFIED': {most_active_unclassified_user}")
            print_to_file(f"Quantidade de commits 'UNCLASSIFIED' realizados pelo usuário: {num_commits_most_active_unclassified_user}")

            unclassified_commits['Date'] = pd.to_datetime(unclassified_commits['Date'])
            unclassified_weekly_commits = unclassified_commits.groupby(unclassified_commits['Date'].dt.day_name()).size().nlargest(3)
            print_to_file("\nTrês dias da semana com mais commits 'UNCLASSIFIED':")
            print_to_file(unclassified_weekly_commits)

            unclassified_commits['Dia_Semana'] = unclassified_commits['Date'].dt.strftime('%A')
            commits_unclassified_por_dia = unclassified_commits['Dia_Semana'].value_counts()
            dia_menos_unclassified_commits = commits_unclassified_por_dia.idxmin()
            num_commits_menos_unclassified = commits_unclassified_por_dia.min()

            print_to_file(f"\nDia com menos commits 'UNCLASSIFIED': {dia_menos_unclassified_commits}")
            print_to_file(f"Quantidade de commits 'UNCLASSIFIED' realizados nesse dia: {num_commits_menos_unclassified}")

            max_files_changed_commit = unclassified_commits.loc[unclassified_commits['Amt_files'].idxmax()]
            user_max_files_changed = max_files_changed_commit['author']
            max_files_changed = max_files_changed_commit['Amt_files']

            print_to_file(f"\nUsuário que fez o commit com maior número de alterações em arquivos: {user_max_files_changed}")
            print_to_file(f"Quantidade de arquivos alterados: {max_files_changed}")

            print_to_file("======================================================================")
            max_files_changed_commit = unclassified_commits.loc[unclassified_commits['Amt_files'].idxmax()]
            user_max_files_changed = max_files_changed_commit['author']
            max_files_changed = max_files_changed_commit['Amt_files']
            commit_message_short = max_files_changed_commit['Short_Message']
            commit_message_full = max_files_changed_commit['Full_Message']
            print_to_file(f"Mensagem do commit (Short): {commit_message_short}")
            print_to_file(f"Mensagem do commit (Full): {commit_message_full}")
        else:
            print_to_file("\nNão há commits 'UNCLASSIFIED' para analisar.")

def data_color_html(nome):
    html_content = f"<html><head><title>Análise do Repositório: {nome}</title></head><body>"

    def add_to_html(content):
        nonlocal html_content
        html_content += content

    add_to_html(f"<h1>Repositório: {user}/{nome}</h1>")

    def get_directory_size(directory):
        total_size = 0
        for dirpath, dirnames, filenames in os.walk(directory):
            for file in filenames:
                fp = os.path.join(dirpath, file)
                if os.path.isfile(fp):
                    total_size += os.path.getsize(fp)
        return total_size

    current_path = os.getcwd()
    repo_path = os.path.join(current_path, f'{nome}')
    size = get_directory_size(repo_path)
    add_to_html(f"<p>Tamanho do repositório: {size / 1024 / 1024:.2f} MB</p>")

    # add link p repo
    repo_link = f"https://github.com/{user}/{nome}"
    add_to_html(f"<p>Link para o repositório GitHub: <a href='{repo_link}' target='_blank'>{repo_link}</a></p>")

    try:
        df = pd.read_excel(f'{nome}_class_color.xlsx')
    except FileNotFoundError:
        add_to_html("<p>Arquivo de dados não encontrado.</p>")
        return html_content + "</body></html>"

    add_to_html("<h2>Linguagens incorporadas:</h2>")
    # Chame a função para obter o conteúdo do diretório clonado
    local_path = os.path.join(current_path, nome)
    if os.path.exists(local_path):
        total_files, extensions = get_local_repository_contents(local_path)

        # linguagens incorporadas no projeto
        languages_incorporated = set()
        for extension, files in extensions.items():
            if extension in extension_to_language:
                languages_incorporated.add(extension_to_language[extension])

        for language in languages_incorporated:
            add_to_html(f"<p>{language}</p>")
    else:
        add_to_html(f"<p>O diretório {local_path} não foi encontrado.</p>")

    commit_counts_per_group = df['Classification'].value_counts()
    add_to_html("<h2>Quantidade de commits por grupo:</h2>")
    add_to_html(commit_counts_per_group.to_frame().to_html())

    total_commits = len(df)
    add_to_html(f"<p>Quantidade total de commits: {total_commits}</p>")

    most_active_user = df['author'].value_counts().idxmax()
    num_commits_most_active_user = df['author'].value_counts().max()
    add_to_html(f"<p>Usuário que mais fez commits: {most_active_user}</p>")
    add_to_html(f"<p>Quantidade de commits realizados pelo usuário: {num_commits_most_active_user}</p>")

    num_users = df['author'].nunique()
    add_to_html(f"<p>Quantidade de usuários: {num_users}</p>")

    df['Date'] = pd.to_datetime(df['Date'])
    df_weekly_commits = df.groupby(df['Date'].dt.strftime('%A')).size().nlargest(3)
    add_to_html("<h2>Três dias da semana com mais commits:</h2>")
    add_to_html(df_weekly_commits.to_frame().to_html())

    df['Dia_Semana'] = df['Date'].dt.strftime('%A')
    commits_por_dia = df['Dia_Semana'].value_counts()
    dia_menos_commits = commits_por_dia.idxmin()
    num_commits_menos = commits_por_dia.min()
    add_to_html(f"<p>Dia do projeto com menos commits: {dia_menos_commits}</p>")
    add_to_html(f"<p>Quantidade de commits realizados nesse dia: {num_commits_menos}</p>")

    add_to_html("<h2>Probabilidade de ocorrência:</h2>")
    total_commits = commit_counts_per_group.sum()
    prob_per_group = commit_counts_per_group / total_commits
    add_to_html(prob_per_group.to_frame().to_html())

    media = prob_per_group.mean().round(20)
    variancia = prob_per_group.var()
    desvio_padrao = prob_per_group.std()

    add_to_html("<h2>Métricas por grupo:</h2>")
    add_to_html(f"<p>Média: {media.round(20)}</p>")
    add_to_html(f"<p>Variância: {variancia}</p>")

    media = prob_per_group.mean()
    desvios_padrao_por_grupo = []

    for group, probabilidade in prob_per_group.items():
        variancia = (probabilidade - media) ** 2
        desvio_padrao_grupo = np.sqrt(variancia)
        desvios_padrao_por_grupo.append((group, desvio_padrao_grupo))

    add_to_html("<h2>Desvio Padrão por grupo:</h2>")
    for grupo, desvio_padrao in desvios_padrao_por_grupo:
        add_to_html(f"<p>{grupo}: {desvio_padrao}</p>")

    add_to_html("<h2>################====UNCLASSIFIED=====####################</h2>")

    unclassified_commits = df[df['Classification'] == 'UNCLASSIFIED'].copy()
    num_unclassified_commits = len(unclassified_commits)
    add_to_html(f"<p>Número de commits 'UNCLASSIFIED': {num_unclassified_commits}</p>")

    if num_unclassified_commits > 0:
        most_active_unclassified_user = unclassified_commits['author'].value_counts().idxmax()
        num_commits_most_active_unclassified_user = unclassified_commits['author'].value_counts().max()

        add_to_html(f"<p>Usuário que mais fez commits 'UNCLASSIFIED': {most_active_unclassified_user}</p>")
        add_to_html(f"<p>Quantidade de commits 'UNCLASSIFIED' realizados pelo usuário: {num_commits_most_active_unclassified_user}</p>")

        unclassified_commits['Date'] = pd.to_datetime(unclassified_commits['Date'])
        unclassified_weekly_commits = unclassified_commits.groupby(unclassified_commits['Date'].dt.day_name()).size().nlargest(3)
        add_to_html("<h2>Três dias da semana com mais commits 'UNCLASSIFIED':</h2>")
        add_to_html(unclassified_weekly_commits.to_frame().to_html())

        unclassified_commits['Dia_Semana'] = unclassified_commits['Date'].dt.strftime('%A')
        commits_unclassified_por_dia = unclassified_commits['Dia_Semana'].value_counts()
        dia_menos_unclassified_commits = commits_unclassified_por_dia.idxmin()
        num_commits_menos_unclassified = commits_unclassified_por_dia.min()

        add_to_html(f"<p>Dia com menos commits 'UNCLASSIFIED': {dia_menos_unclassified_commits}</p>")
        add_to_html(f"<p>Quantidade de commits 'UNCLASSIFIED' realizados nesse dia: {num_commits_menos_unclassified}</p>")

        max_files_changed_commit = unclassified_commits.loc[unclassified_commits['Amt_files'].idxmax()]
        user_max_files_changed = max_files_changed_commit['author']
        max_files_changed = max_files_changed_commit['Amt_files']

        add_to_html(f"<p>Usuário que fez o commit com maior número de alterações em arquivos: {user_max_files_changed}</p>")
        add_to_html(f"<p>Quantidade de arquivos alterados: {max_files_changed}</p>")

        commit_message_short = max_files_changed_commit['Short_Message']
        commit_message_full = max_files_changed_commit['Full_Message']
        add_to_html("<p>======================================================================</p>")
        add_to_html(f"<p>Mensagem do commit (Short): {commit_message_short}</p>")
        add_to_html(f"<p>Mensagem do commit (Full): {commit_message_full}</p>")
    else:
        add_to_html("<p>Não há commits 'UNCLASSIFIED' para analisar.</p>")

    add_to_html('</body></html>')

    html_file = f'{nome}_output.html'
    with open(html_file, 'w', encoding='utf-8') as file:
        file.write(html_content)



def graph():
  arquivo = f'{nome}_class_color.xlsx'
  df = pd.read_excel(arquivo)
  df['Full_Message'].fillna('', inplace=True)

  cores = {
      'Forward Engineering': 'blue',
      'Reengineering': 'lime',
      'Corrective Engineering': 'orange',
      'Management': 'brown',
      'Tests': 'yellow',
      'UNCLASSIFIED': 'gray'
  }

  df['Hour'] = pd.to_timedelta(df['Hour'])
  df_sorted = df.sort_values(by='Hour')
  fig = px.scatter(df_sorted, x='Date', y='Hour', color='Classification', color_discrete_map=cores,
                  title='Horário dos Commits por Categoria', hover_data={'Short_Message': True, 'Full_Message': True})
  fig.update_layout(xaxis_title='Date', yaxis_title='Hora do Commit')
  fig.update_xaxes(tickangle=45)
  fig.show()



def download():
    folder_path = '/content'
    files_in_folder = os.listdir(folder_path)
    for file in files_in_folder:
        if file.endswith('.xlsx') or file.endswith('.txt') or file.endswith('.html') or file.endswith('.py'):
            files.download(os.path.join(folder_path, file))

def clear():
  pasta_principal = '/content'
  for arquivo in os.listdir(pasta_principal):
      caminho_completo = os.path.join(pasta_principal, arquivo)
      if os.path.isfile(caminho_completo):
          os.remove(caminho_completo)

def gerar_pasta(nome):
    script_content = f"""
import os

def criar_pasta():
    # Nome da pasta a ser criada
    nome_pasta = '{nome}'

    # Caminho completo para a nova pasta
    nova_pasta_path = os.path.join(os.getcwd(), nome_pasta)

    # Cria a pasta se ela não existir
    if not os.path.exists(nova_pasta_path):
        os.makedirs(nova_pasta_path)
        print(f"Pasta '{nome}' criada com sucesso em {os.getcwd()}.")
    else:
        print(f"A pasta '{nome}' já existe em {os.getcwd()}.")

if __name__ == "__main__":
    criar_pasta()
    """

    nome_script = f'gerar_pasta_{nome}.py'
    with open(nome_script, 'w') as f:
        f.write(script_content)

    print(f"Script '{nome_script}' gerado com sucesso!")


def lang():
  local_path = f'/content/{nome}'  # Ajuste este caminho conforme necessário
  try:
      if os.path.exists(local_path):
          print(f"Atualizando repositório em {local_path}...")
          repo = Repo(local_path)
          repo.remotes.origin.pull()
      else:
          print(f"O diretório {local_path} não existe. Certifique-se de que o repositório está clonado.")
  except GitCommandError as e:
      print(f"Erro ao atualizar o repositório: {e}")

  def get_local_repository_contents2(directory, file_count=0, extensions=None):
      if extensions is None:
          extensions = {}

      for dirpath, dirnames, filenames in os.walk(directory):
          for filename in filenames:
              file_count += 1
              file_path = os.path.join(dirpath, filename)
              # Adicionar saída para o arquivo de texto
              with open('languages.txt', 'a') as f:
                  print(f"Arquivo: {file_path}", file=f)
              # Obter a extensão do arquivo
              file_extension = filename.split('.')[-1]
              # Atualizar o dicionário de extensões
              extensions[file_extension] = extensions.get(file_extension, []) + [file_path]

      return file_count, extensions

  if os.path.exists(local_path):
      total_files, extensions = get_local_repository_contents2(local_path)
      # Salvar o total de arquivos no arquivo de texto
      with open('languages.txt', 'a') as f:
          print(f"Total de arquivos: {total_files}", file=f)

      languages_incorporated = set()
      for extension, files in extensions.items():
          if extension in extension_to_language:
              languages_incorporated.add(extension_to_language[extension])

      with open('languages.txt', 'a') as f:
          print("Linguagens incorporadas no projeto:", file=f)
          for language in languages_incorporated:
              print(language, file=f)
  else:
      print(f"O diretório {local_path} não foi encontrado.")

def gitgraph():
  saida = ".."
  %cd "{nome}"
  !git graph
  %cd "{saida}"
  # Localizar o arquivo PDF gerado
  pdf_path = f'/content/{nome}/.gitGraph/*.pdf'
  generated_pdf = !ls -tr {pdf_path} | tail -n 1

  # Renomear o arquivo PDF para facilitar o download
  pdf_file = generated_pdf[0]
  pdf_basename = os.path.basename(pdf_file)
  pdf_download_path = f'/content/{pdf_basename}'
  shutil.move(pdf_file, pdf_download_path)

  # Baixar o arquivo PDF
  files.download(pdf_download_path)

def gitgraph2(nome):
    try:
        saida = ".."  # Diretório de saída

        # Mudar para o diretório do repositório
        get_ipython().run_line_magic('cd', f'"{nome}"')

        # Executar o comando git-graph para gerar o PDF
        !git graph -n "{nome}_graph" -f pdf

        # Voltar para o diretório de saída
        get_ipython().run_line_magic('cd', f'"{saida}"')

        # Localizar o arquivo PDF gerado
        pdf_path = f'/content/{nome}_graph.pdf'

        # Verificar se o arquivo PDF existe
        if os.path.exists(pdf_path):
            # Renomear o arquivo PDF para facilitar o download
            pdf_basename = f"{nome}_graph.pdf"
            pdf_download_path = f'/content/{pdf_basename}'
            shutil.move(pdf_path, pdf_download_path)

            # Baixar o arquivo PDF
            from google.colab import files
            files.download(pdf_download_path)
        else:
            print('Nenhum arquivo PDF encontrado.')

    except Exception as e:
        print(f'Ocorreu um erro: {e}')

def move():
    entrada = f'{nome}/.gitGraph'  # Nome da pasta oculta
    saida = '..'  # Pasta raiz

    # Mudar para a pasta oculta
    %cd "{entrada}"

    # Listar todos os arquivos na pasta oculta
    files = os.listdir()

    # Mover cada arquivo para a pasta raiz
    for file in files:
        shutil.move(file, f"{saida}/{file}")

    # Voltar para a pasta raiz
    %cd "{saida}"
    %cd "{saida}"

##################### fixos

def get_local_repository_contents(directory, file_count=0, extensions=None):
    if extensions is None:
        extensions = {}

    for dirpath, dirnames, filenames in os.walk(directory):
        for filename in filenames:
            file_count += 1
            file_path = os.path.join(dirpath, filename)
            # Obter a extensão do arquivo
            file_extension = filename.split('.')[-1]
            # Atualizar o dicionário de extensões
            extensions[file_extension] = extensions.get(file_extension, []) + [file_path]

    return file_count, extensions

